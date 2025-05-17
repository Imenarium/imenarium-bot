import os
import re
import logging
from aiogram import Bot, Dispatcher, executor, types
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InputFile
from openpyxl import load_workbook
from datetime import datetime

# === НАСТРОЙКИ ===
API_TOKEN = os.getenv("BOT_TOKEN")  # Установи переменную окружения или вставь токен напрямую
SCENARIO_PATH = "scenario.xlsx"
MEDIA_FOLDER = "data"

# === НАСТРОЙКА ЛОГИРОВАНИЯ ===
logging.basicConfig(level=logging.INFO)
bot = Bot(token=API_TOKEN)
dp = Dispatcher(bot)

# === МЕНЮ ===
main_menu = ReplyKeyboardMarkup(resize_keyboard=True)
main_menu.add(
    KeyboardButton("В начало"),
    KeyboardButton("Как заказать книгу")
).add(
    KeyboardButton("Доставка и оплата"),
    KeyboardButton("В чат с оператором")
).add(
    KeyboardButton("Перейти на сайт")
)

# === ЗАГРУЗКА СЦЕНАРИЯ ===
SCENARIO = {}
KEYWORDS = {}

def load_scenario():
    wb = load_workbook(SCENARIO_PATH)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        block_id, text, buttons, transitions, keywords, media = row
        block = {
            "text": text,
            "buttons": [b.strip() for b in str(buttons).split('|')] if buttons else [],
            "transitions": [t.strip() for t in str(transitions).split('|')] if transitions else [],
            "media": media
        }
        SCENARIO[block_id] = block

        if keywords:
            for kw in str(keywords).split(','):
                KEYWORDS[kw.strip().lower()] = block_id

# === ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ===
def get_greeting():
    hour = datetime.now().hour
    if 5 <= hour < 12:
        return "Доброе утро"
    elif 12 <= hour < 18:
        return "Добрый день"
    elif 18 <= hour < 24:
        return "Добрый вечер"
    else:
        return "Доброй ночи"

def render_block(block_id, message):
    if block_id not in SCENARIO:
        return message.answer("Блок не найден :(")
    block = SCENARIO[block_id]
    text = block["text"].replace("{имя}", message.from_user.first_name)
    keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
    for btn in block["buttons"]:
        keyboard.add(KeyboardButton(btn))

    if block["media"]:
        file_path = os.path.join(MEDIA_FOLDER, block["media"])
        if os.path.exists(file_path):
            return message.answer_photo(InputFile(file_path), caption=text, reply_markup=keyboard)
    return message.answer(text, reply_markup=keyboard)

# === СТАРТ ===
@dp.message_handler(commands=['start'])
async def send_welcome(message: types.Message):
    greet = get_greeting()
    await render_block("Приветствие", message)

# === ОБРАБОТКА КНОПОК ===
@dp.message_handler(lambda message: message.text in SCENARIO)
async def handle_exact_block(message: types.Message):
    await render_block(message.text, message)

# === ОБРАБОТКА КЛЮЧЕВЫХ СЛОВ ===
@dp.message_handler(lambda message: message.text)
async def handle_keywords(message: types.Message):
    text = message.text.lower()
    for kw, block_id in KEYWORDS.items():
        if re.search(kw, text):
            return await render_block(block_id, message)

    await message.answer(f"{message.from_user.first_name}, если вы не нашли нужную кнопку — я с радостью помогу! Просто выберите пункт внизу 👇", reply_markup=main_menu)

# === ЗАПУСК ===
if __name__ == '__main__':
    load_scenario()
    executor.start_polling(dp, skip_updates=True)
